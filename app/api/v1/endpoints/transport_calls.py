from __future__ import annotations

import csv
import io
import unicodedata
import uuid
from datetime import datetime
from typing import Any
from urllib.parse import parse_qs

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, Request, Response, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import ADMIN_ROLES, READ_ROLES, WRITE_ROLES, get_db, require_company_user_for_tenant, require_enabled_module
from app.web.admin_v2_routes import _active_session as active_admin_v2_session
from app.web.admin_v2_routes import _active_company_preview as active_admin_company_preview

router = APIRouter()

TRANSPORT_CALL_MANAGER_ROLES = ADMIN_ROLES | {"manager", "gerencia", "gerente", "supervisor", "tesoreria"}
TRANSPORT_CALL_IMPORT_MAX_BYTES = 8 * 1024 * 1024
TRANSPORT_CALL_IMPORT_MAX_ROWS = 10_000


def _authorization_from_query_028s(authorization: str | None, access_token: str | None = "") -> str | None:
    token = str(access_token or "").strip()
    return authorization or (f"Bearer {token}" if token else None)


async def require_transport_calls_read(
    company_id: uuid.UUID,
    request: Request,
    authorization: str | None = Header(default=None),
    access_token: str = Query(default="", max_length=2048),
    db: AsyncSession = Depends(get_db),
) -> None:
    if await active_admin_v2_session(request, db) or active_admin_company_preview(request, company_id):
        await require_enabled_module(db, company_id, "transport_calls")
        return
    await require_company_user_for_tenant(
        db,
        _authorization_from_query_028s(authorization, access_token),
        company_id,
        allowed_roles=READ_ROLES,
        module_codes="transport_calls",
    )


async def require_transport_calls_write(
    company_id: uuid.UUID,
    request: Request,
    authorization: str | None = Header(default=None),
    db: AsyncSession = Depends(get_db),
) -> None:
    if await active_admin_v2_session(request, db) or active_admin_company_preview(request, company_id):
        await require_enabled_module(db, company_id, "transport_calls")
        return
    await require_company_user_for_tenant(
        db,
        authorization,
        company_id,
        allowed_roles=WRITE_ROLES,
        module_codes="transport_calls",
    )


async def require_transport_calls_manage(
    company_id: uuid.UUID,
    request: Request,
    authorization: str | None = Header(default=None),
    access_token: str = Query(default="", max_length=2048),
    db: AsyncSession = Depends(get_db),
) -> None:
    if await active_admin_v2_session(request, db) or active_admin_company_preview(request, company_id):
        await require_enabled_module(db, company_id, "transport_calls")
        return
    await require_company_user_for_tenant(
        db,
        _authorization_from_query_028s(authorization, access_token),
        company_id,
        allowed_roles=TRANSPORT_CALL_MANAGER_ROLES,
        module_codes="transport_calls",
    )


class TransportCallIn(BaseModel):
    advisor_name: str | None = Field(default="", max_length=180)
    advisor_status: str | None = Field(default="available", max_length=40)
    customer_name: str | None = Field(default="", max_length=180)
    customer_type: str | None = Field(default="person", max_length=40)
    phone: str | None = Field(default="", max_length=80)
    origin: str | None = Field(default="", max_length=160)
    destination: str | None = Field(default="", max_length=160)
    trip_type: str | None = Field(default="", max_length=80)
    call_direction: str | None = Field(default="inbound", max_length=20)
    call_status: str | None = Field(default="completed", max_length=40)
    result: str | None = Field(default="follow_up", max_length=80)
    duration_seconds: int | None = Field(default=0, ge=0)
    duration_minutes: float | None = Field(default=None, ge=0)
    quote_requested: bool | None = False
    ticket_requested: bool | None = False
    contract_code: str | None = Field(default="", max_length=120)
    source: str | None = Field(default="manual", max_length=40)
    twilio_call_sid: str | None = Field(default="", max_length=120)
    twilio_parent_call_sid: str | None = Field(default="", max_length=120)
    batch_row_id: str | None = Field(default="", max_length=80)
    campaign_code: str | None = Field(default="", max_length=120)
    caller_number: str | None = Field(default="", max_length=80)
    phone_type: str | None = Field(default="unknown", max_length=30)
    price_amount: float | None = Field(default=0, ge=0)
    price_currency: str | None = Field(default="USD", max_length=12)
    consent_status: str | None = Field(default="unknown", max_length=30)
    do_not_call: bool | None = False
    notes: str | None = Field(default="", max_length=1200)


def _clean(value: Any, limit: int = 255) -> str:
    return str(value or "").strip()[:limit]


def _optional_uuid(value: Any, detail: str = "id_invalid") -> str:
    raw = _clean(value, 80)
    if not raw:
        return ""
    try:
        return str(uuid.UUID(raw))
    except (TypeError, ValueError, AttributeError) as exc:
        raise HTTPException(status_code=400, detail=detail) from exc


def _csv_key(value: Any) -> str:
    return (
        unicodedata.normalize("NFKD", str(value or ""))
        .encode("ascii", "ignore")
        .decode("ascii")
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("/", "_")
        .replace("-", "_")
    )


def _csv_value(row: dict[str, Any], *names: str, limit: int = 255) -> str:
    normalized = {_csv_key(key): value for key, value in row.items()}
    for name in names:
        key = _csv_key(name)
        if key in normalized and normalized[key] not in (None, ""):
            return _clean(normalized[key], limit)
    return ""


def _csv_money(row: dict[str, Any], *names: str) -> float:
    raw = _csv_value(row, *names, limit=80)
    if not raw:
        return 0.0
    clean = "".join(ch for ch in raw if ch.isdigit() or ch in ",.-")
    if not clean:
        return 0.0
    try:
        if "," in clean and "." in clean:
            clean = clean.replace(".", "").replace(",", ".") if clean.rfind(",") > clean.rfind(".") else clean.replace(",", "")
        elif clean.count(".") == 1 and len(clean.rsplit(".", 1)[-1]) == 3:
            clean = clean.replace(".", "")
        elif clean.count(",") == 1 and len(clean.rsplit(",", 1)[-1]) == 3:
            clean = clean.replace(",", "")
        elif clean.count(".") > 1:
            clean = clean.replace(".", "")
        elif clean.count(",") > 1:
            clean = clean.replace(",", "")
        else:
            clean = clean.replace(",", ".")
        return max(0.0, float(clean or 0))
    except ValueError:
        return 0.0


def _csv_bool(row: dict[str, Any], *names: str) -> bool:
    return _csv_value(row, *names, limit=30).lower() in {"1", "true", "si", "yes", "x"}


def _tabular_call_rows(source_rows: Any) -> list[dict[str, Any]]:
    phone_headers = {"telefono", "phone", "celular", "whatsapp"}
    identity_headers = {
        "cliente",
        "customer_name",
        "nombre",
        "empresa",
        "razon_social",
        "contrato",
        "aval",
        "contrato_aval",
    }
    known_headers = phone_headers | identity_headers | {
        "tipo_cliente",
        "correo",
        "email",
        "documento",
        "cuenta",
        "origen",
        "destino",
        "tipo_viaje",
        "transportadora",
        "valor_ticket",
        "campana",
        "tipo_linea",
        "consentimiento",
        "no_llamar",
        "notas",
    }
    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    for source_row in source_rows:
        values = list(source_row or [])
        if not headers:
            normalized = {_csv_key(value) for value in values if value not in (None, "")}
            if normalized & identity_headers and len(normalized & known_headers) >= 2:
                headers = [str(value or "").strip() for value in values]
            continue
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({header: value for header, value in zip(headers, values) if header})
        if len(rows) > TRANSPORT_CALL_IMPORT_MAX_ROWS:
            raise ValueError(f"La base supera el maximo de {TRANSPORT_CALL_IMPORT_MAX_ROWS} registros.")
    if not headers:
        raise ValueError("No se encontro la fila de encabezados con cliente y telefono.")
    return rows


def _csv_rows(body: str) -> list[dict[str, Any]]:
    try:
        dialect = csv.Sniffer().sniff(body[:8192], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    try:
        rows = _tabular_call_rows(csv.reader(io.StringIO(body, newline=""), dialect=dialect))
    except csv.Error as exc:
        raise ValueError("El CSV no se pudo leer. Exportalo nuevamente como CSV UTF-8.") from exc
    return rows


def _xlsx_rows(raw: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("El archivo Excel no se pudo leer. Usa un archivo .xlsx valido.") from exc
    try:
        try:
            worksheets = [workbook.active, *(sheet for sheet in workbook.worksheets if sheet is not workbook.active)]
            header_error: ValueError | None = None
            for worksheet in worksheets:
                try:
                    rows = _tabular_call_rows(worksheet.iter_rows(values_only=True))
                except ValueError as exc:
                    if "encabezados" not in str(exc):
                        raise
                    header_error = exc
                    continue
                if rows:
                    return rows
            if header_error:
                raise header_error
            return []
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError("El archivo Excel no se pudo leer. Usa un archivo .xlsx valido.") from exc
    finally:
        workbook.close()


def _uploaded_call_rows(raw: bytes, filename: str) -> list[dict[str, Any]]:
    if not raw:
        raise ValueError("El archivo esta vacio.")
    if len(raw) > TRANSPORT_CALL_IMPORT_MAX_BYTES:
        raise ValueError("El archivo supera el limite de 8 MB.")
    lower_name = str(filename or "").strip().lower()
    if lower_name.endswith(".xlsx") or raw.startswith(b"PK\x03\x04"):
        rows = _xlsx_rows(raw)
    elif lower_name.endswith((".csv", ".txt")) or "." not in lower_name:
        try:
            body = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            body = raw.decode("latin-1")
        rows = _csv_rows(body)
        if len(rows) > TRANSPORT_CALL_IMPORT_MAX_ROWS:
            raise ValueError(f"La base supera el maximo de {TRANSPORT_CALL_IMPORT_MAX_ROWS} registros.")
    else:
        raise ValueError("Formato no compatible. Selecciona un archivo .xlsx o .csv.")
    if not rows:
        raise ValueError("La base no contiene registros.")
    if not any(_csv_value(row, "telefono", "phone", "celular", "whatsapp") for row in rows):
        raise ValueError("La base debe incluir la columna telefono y al menos un numero.")
    return rows


def _role_tokens(*values: Any) -> set[str]:
    raw = " ".join(str(value or "") for value in values)
    normalized = (
        unicodedata.normalize("NFKD", raw)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
        .replace("/", " ")
        .replace("-", " ")
        .replace("_", " ")
    )
    compact = normalized.replace(" ", "")
    return set(normalized.split()) | ({compact} if compact else set())


def _is_call_agent_role(role: Any, employee_type: Any = "", panel_type: Any = "") -> bool:
    tokens = _role_tokens(role, employee_type, panel_type)
    allowed = {
        "agente", "asesor", "operario", "vendedor", "agente_call", "agentecall",
        "agente_call_center", "agentecallcenter", "call", "callcenter", "call_center",
        "asesorcall", "asesorcallcenter", "agenteexterno", "externo",
    }
    return bool(tokens & allowed)


def _duration_seconds(payload: TransportCallIn) -> int:
    if payload.duration_minutes is not None:
        return max(0, int(float(payload.duration_minutes or 0) * 60))
    return max(0, int(payload.duration_seconds or 0))


def _row(row: Any) -> dict[str, Any]:
    data = dict(row._mapping if hasattr(row, "_mapping") else row)
    for key, value in list(data.items()):
        if isinstance(value, (datetime,)):
            data[key] = value.isoformat()
        elif isinstance(value, uuid.UUID):
            data[key] = str(value)
    return data


async def ensure_transport_calls_storage(db: AsyncSession) -> None:
    await db.execute(text('CREATE EXTENSION IF NOT EXISTS "pgcrypto";'))
    await db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS transport_call_logs (
                id uuid PRIMARY KEY DEFAULT gen_random_uuid(),
                company_id uuid NOT NULL,
                advisor_name varchar(180) NOT NULL DEFAULT '',
                advisor_status varchar(40) NOT NULL DEFAULT 'available',
                customer_name varchar(180) NOT NULL DEFAULT '',
                customer_type varchar(40) NOT NULL DEFAULT 'person',
                phone varchar(80) NOT NULL DEFAULT '',
                origin varchar(160) NOT NULL DEFAULT '',
                destination varchar(160) NOT NULL DEFAULT '',
                trip_type varchar(80) NOT NULL DEFAULT '',
                call_direction varchar(20) NOT NULL DEFAULT 'inbound',
                call_status varchar(40) NOT NULL DEFAULT 'completed',
                result varchar(80) NOT NULL DEFAULT 'follow_up',
                duration_seconds integer NOT NULL DEFAULT 0,
                quote_requested boolean NOT NULL DEFAULT false,
                ticket_requested boolean NOT NULL DEFAULT false,
                contract_code varchar(120) NOT NULL DEFAULT '',
                notes text NOT NULL DEFAULT '',
                created_at timestamptz NOT NULL DEFAULT now(),
                updated_at timestamptz NOT NULL DEFAULT now()
            )
            """
        )
    )
    for statement in [
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS advisor_status varchar(40) NOT NULL DEFAULT 'available'",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS contract_code varchar(120) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS source varchar(40) NOT NULL DEFAULT 'manual'",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS twilio_call_sid varchar(120) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS twilio_parent_call_sid varchar(120) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS batch_row_id UUID NULL",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS campaign_code varchar(120) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS caller_number varchar(80) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS phone_type varchar(30) NOT NULL DEFAULT 'unknown'",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS price_amount numeric(14,6) NOT NULL DEFAULT 0",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS price_currency varchar(12) NOT NULL DEFAULT 'USD'",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS consent_status varchar(30) NOT NULL DEFAULT 'unknown'",
        "ALTER TABLE transport_call_logs ADD COLUMN IF NOT EXISTS do_not_call boolean NOT NULL DEFAULT false",
        "CREATE INDEX IF NOT EXISTS ix_transport_call_logs_company_created ON transport_call_logs (company_id, created_at DESC)",
        "CREATE INDEX IF NOT EXISTS ix_transport_call_logs_company_advisor ON transport_call_logs (company_id, advisor_name)",
        "CREATE INDEX IF NOT EXISTS ix_transport_call_logs_company_twilio ON transport_call_logs (company_id, twilio_call_sid)",
    ]:
        await db.execute(text(statement))

    await db.execute(text("""
        CREATE TABLE IF NOT EXISTS transport_call_batches (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            company_id UUID NOT NULL,
            file_name VARCHAR(240) NOT NULL DEFAULT '',
            assigned_employee_id UUID NULL,
            assigned_agent_name VARCHAR(180) NOT NULL DEFAULT '',
            assigned_agent_role VARCHAR(80) NOT NULL DEFAULT '',
            assigned_user_id UUID NULL,
            record_count INTEGER NOT NULL DEFAULT 0,
            status VARCHAR(40) NOT NULL DEFAULT 'active',
            created_by_label VARCHAR(180) NOT NULL DEFAULT '',
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            archived_at TIMESTAMPTZ NULL
        )
    """))
    await db.execute(text("""
        CREATE TABLE IF NOT EXISTS transport_call_batch_rows (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            batch_id UUID NOT NULL REFERENCES transport_call_batches(id) ON DELETE CASCADE,
            company_id UUID NOT NULL,
            row_number INTEGER NOT NULL DEFAULT 0,
            customer_name VARCHAR(180) NOT NULL DEFAULT '',
            customer_type VARCHAR(40) NOT NULL DEFAULT 'person',
            phone VARCHAR(80) NOT NULL DEFAULT '',
            email VARCHAR(160) NOT NULL DEFAULT '',
            document_id VARCHAR(80) NOT NULL DEFAULT '',
            contract_code VARCHAR(120) NOT NULL DEFAULT '',
            account_code VARCHAR(120) NOT NULL DEFAULT '',
            origin VARCHAR(160) NOT NULL DEFAULT '',
            destination VARCHAR(160) NOT NULL DEFAULT '',
            trip_type VARCHAR(80) NOT NULL DEFAULT '',
            transporter VARCHAR(180) NOT NULL DEFAULT '',
            ticket_value NUMERIC(14,2) NOT NULL DEFAULT 0,
            call_direction VARCHAR(20) NOT NULL DEFAULT '',
            call_status VARCHAR(40) NOT NULL DEFAULT '',
            result VARCHAR(80) NOT NULL DEFAULT '',
            duration_seconds INTEGER NOT NULL DEFAULT 0,
            quote_requested BOOLEAN NOT NULL DEFAULT false,
            ticket_requested BOOLEAN NOT NULL DEFAULT false,
            twilio_call_sid VARCHAR(120) NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            managed_by VARCHAR(180) NOT NULL DEFAULT '',
            managed_at TIMESTAMPTZ NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
    """))
    for statement in [
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS account_code VARCHAR(120) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS transporter VARCHAR(180) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS ticket_value NUMERIC(14,2) NOT NULL DEFAULT 0",
        "ALTER TABLE transport_call_batches ADD COLUMN IF NOT EXISTS campaign_code VARCHAR(120) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS campaign_code VARCHAR(120) NOT NULL DEFAULT ''",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS phone_type VARCHAR(30) NOT NULL DEFAULT 'unknown'",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS consent_status VARCHAR(30) NOT NULL DEFAULT 'unknown'",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS do_not_call BOOLEAN NOT NULL DEFAULT false",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS price_amount NUMERIC(14,6) NOT NULL DEFAULT 0",
        "ALTER TABLE transport_call_batch_rows ADD COLUMN IF NOT EXISTS price_currency VARCHAR(12) NOT NULL DEFAULT 'USD'",
        "CREATE INDEX IF NOT EXISTS ix_transport_call_batches_company ON transport_call_batches(company_id, status)",
        "CREATE INDEX IF NOT EXISTS ix_transport_call_batches_agent ON transport_call_batches(company_id, assigned_employee_id, status)",
        "CREATE INDEX IF NOT EXISTS ix_transport_call_batch_rows_batch ON transport_call_batch_rows(batch_id, row_number)",
        "CREATE INDEX IF NOT EXISTS ix_transport_call_batch_rows_company ON transport_call_batch_rows(company_id, managed_at)",
    ]:
        await db.execute(text(statement))


@router.get("/companies/{company_id}/calls", dependencies=[Depends(require_transport_calls_read)])
async def list_transport_calls(
    company_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    limit: int = Query(default=80, ge=1, le=250),
    search: str = Query(default="", max_length=120),
    start_date: str = Query(default="", max_length=20),
    end_date: str = Query(default="", max_length=20),
) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    query = f"%{_clean(search, 120).lower()}%"
    result = await db.execute(
        text(
            """
            SELECT *
            FROM transport_call_logs
            WHERE company_id = :company_id
              AND (:start_date = '' OR created_at >= CAST(:start_date AS date))
              AND (:end_date = '' OR created_at < (CAST(:end_date AS date) + INTERVAL '1 day'))
              AND (
                :query = '%%'
                OR LOWER(advisor_name) LIKE :query
                OR LOWER(customer_name) LIKE :query
                OR LOWER(phone) LIKE :query
                OR LOWER(contract_code) LIKE :query
                OR LOWER(twilio_call_sid) LIKE :query
                OR LOWER(campaign_code) LIKE :query
                OR LOWER(origin) LIKE :query
                OR LOWER(destination) LIKE :query
                OR LOWER(result) LIKE :query
                OR LOWER(notes) LIKE :query
              )
            ORDER BY created_at DESC
            LIMIT :limit
            """
        ),
        {
            "company_id": str(company_id),
            "query": query,
            "limit": int(limit),
            "start_date": _clean(start_date, 20),
            "end_date": _clean(end_date, 20),
        },
    )
    rows = [_row(row) for row in result.fetchall()]
    return {"ok": True, "company_id": str(company_id), "calls": rows, "count": len(rows)}


@router.get("/companies/{company_id}/summary", dependencies=[Depends(require_transport_calls_read)])
async def transport_calls_summary(
    company_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    search: str = Query(default="", max_length=120),
    start_date: str = Query(default="", max_length=20),
    end_date: str = Query(default="", max_length=20),
) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    query = f"%{_clean(search, 120).lower()}%"
    result = await db.execute(
        text(
            """
            WITH period AS (
                SELECT *
                FROM transport_call_logs
                WHERE company_id = :company_id
                  AND (:start_date = '' OR created_at >= CAST(:start_date AS date))
                  AND (:end_date = '' OR created_at < (CAST(:end_date AS date) + INTERVAL '1 day'))
                  AND (
                    :query = '%%'
                    OR LOWER(advisor_name) LIKE :query
                    OR LOWER(customer_name) LIKE :query
                    OR LOWER(phone) LIKE :query
                    OR LOWER(contract_code) LIKE :query
                    OR LOWER(origin) LIKE :query
                    OR LOWER(destination) LIKE :query
                    OR LOWER(result) LIKE :query
                    OR LOWER(notes) LIKE :query
                  )
            ),
            today AS (
                SELECT *
                FROM transport_call_logs
                WHERE company_id = :company_id
                  AND created_at >= date_trunc('day', now())
            ),
            latest_advisor AS (
                SELECT DISTINCT ON (LOWER(advisor_name))
                    advisor_name,
                    advisor_status,
                    created_at
                FROM transport_call_logs
                WHERE company_id = :company_id
                  AND TRIM(advisor_name) <> ''
                ORDER BY LOWER(advisor_name), created_at DESC
            )
            SELECT
                (SELECT COUNT(*) FROM today) AS calls_today,
                (SELECT COUNT(*) FROM transport_call_logs WHERE company_id = :company_id) AS calls_total,
                COALESCE((SELECT SUM(duration_seconds) FROM today), 0) AS duration_today,
                COALESCE((SELECT ROUND(AVG(NULLIF(duration_seconds, 0)))::integer FROM today), 0) AS avg_duration_today,
                (SELECT COUNT(*) FROM today WHERE quote_requested IS TRUE) AS quotes_today,
                (SELECT COUNT(*) FROM today WHERE ticket_requested IS TRUE) AS tickets_today,
                (SELECT COUNT(*) FROM period) AS calls_period,
                COALESCE((SELECT SUM(duration_seconds) FROM period), 0) AS duration_period,
                COALESCE((SELECT ROUND(AVG(NULLIF(duration_seconds, 0)))::integer FROM period), 0) AS avg_duration_period,
                (SELECT COUNT(*) FROM period WHERE quote_requested IS TRUE) AS quotes_period,
                (SELECT COUNT(*) FROM period WHERE ticket_requested IS TRUE) AS tickets_period,
                COALESCE((SELECT SUM(price_amount) FROM period), 0) AS cost_period,
                (SELECT COUNT(*) FROM latest_advisor) AS advisors_total,
                (SELECT COUNT(*) FROM latest_advisor WHERE advisor_status = 'available') AS advisors_available,
                (SELECT COUNT(*) FROM latest_advisor WHERE advisor_status = 'in_call') AS advisors_in_call,
                (SELECT COUNT(*) FROM latest_advisor WHERE advisor_status IN ('break', 'bathroom', 'lunch')) AS advisors_paused,
                (SELECT COUNT(*) FROM today WHERE call_status = 'missed') AS missed_today,
                (SELECT COUNT(*) FROM period WHERE call_status IN ('missed','no_answer','failed')) AS missed_period
            """
        ),
        {
            "company_id": str(company_id),
            "query": query,
            "start_date": _clean(start_date, 20),
            "end_date": _clean(end_date, 20),
        },
    )
    summary = _row(result.first() or {})
    live_agents = (await list_transport_call_agents(company_id, db)).get("agents") or []
    summary.update({
        "advisors_total": len(live_agents),
        "advisors_available": sum(1 for agent in live_agents if agent.get("live_status") == "available"),
        "advisors_in_call": sum(1 for agent in live_agents if agent.get("live_status") == "in_call"),
        "advisors_paused": sum(1 for agent in live_agents if agent.get("live_status") == "break"),
    })
    return {"ok": True, "company_id": str(company_id), "summary": summary}


@router.post("/companies/{company_id}/calls", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_transport_calls_write)])
async def create_transport_call(
    company_id: uuid.UUID,
    payload: TransportCallIn,
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    params = {
        "company_id": str(company_id),
        "advisor_name": _clean(payload.advisor_name, 180),
        "advisor_status": _clean(payload.advisor_status or "available", 40),
        "customer_name": _clean(payload.customer_name, 180),
        "customer_type": _clean(payload.customer_type or "person", 40),
        "phone": _clean(payload.phone, 80),
        "origin": _clean(payload.origin, 160),
        "destination": _clean(payload.destination, 160),
        "trip_type": _clean(payload.trip_type, 80),
        "call_direction": _clean(payload.call_direction or "inbound", 20),
        "call_status": _clean(payload.call_status or "completed", 40),
        "result": _clean(payload.result or "follow_up", 80),
        "duration_seconds": _duration_seconds(payload),
        "quote_requested": bool(payload.quote_requested),
        "ticket_requested": bool(payload.ticket_requested),
        "contract_code": _clean(payload.contract_code, 120),
        "source": _clean(payload.source or "manual", 40),
        "twilio_call_sid": _clean(payload.twilio_call_sid, 120),
        "twilio_parent_call_sid": _clean(payload.twilio_parent_call_sid, 120),
        "batch_row_id": _optional_uuid(payload.batch_row_id, "batch_row_id_invalid"),
        "campaign_code": _clean(payload.campaign_code, 120),
        "caller_number": _clean(payload.caller_number, 80),
        "phone_type": _clean(payload.phone_type or "unknown", 30),
        "price_amount": float(payload.price_amount or 0),
        "price_currency": _clean(payload.price_currency or "USD", 12).upper(),
        "consent_status": _clean(payload.consent_status or "unknown", 30),
        "do_not_call": bool(payload.do_not_call),
        "notes": _clean(payload.notes, 1200),
    }
    row = None
    call_sid = params["twilio_call_sid"] or params["twilio_parent_call_sid"]
    if call_sid:
        existing = await db.execute(text("""
            UPDATE transport_call_logs
            SET advisor_name = COALESCE(NULLIF(:advisor_name, ''), advisor_name),
                advisor_status = :advisor_status,
                customer_name = COALESCE(NULLIF(:customer_name, ''), customer_name),
                customer_type = :customer_type,
                phone = COALESCE(NULLIF(:phone, ''), phone),
                origin = :origin, destination = :destination, trip_type = :trip_type,
                call_direction = :call_direction, call_status = :call_status, result = :result,
                duration_seconds = GREATEST(duration_seconds, :duration_seconds),
                quote_requested = :quote_requested, ticket_requested = :ticket_requested,
                contract_code = :contract_code, source = :source,
                twilio_call_sid = COALESCE(NULLIF(:twilio_call_sid, ''), twilio_call_sid),
                twilio_parent_call_sid = COALESCE(NULLIF(:twilio_parent_call_sid, ''), twilio_parent_call_sid),
                batch_row_id = COALESCE(CAST(NULLIF(:batch_row_id, '') AS uuid), batch_row_id),
                campaign_code = COALESCE(NULLIF(:campaign_code, ''), campaign_code),
                caller_number = COALESCE(NULLIF(:caller_number, ''), caller_number),
                phone_type = :phone_type,
                price_amount = GREATEST(price_amount, :price_amount), price_currency = :price_currency,
                consent_status = :consent_status, do_not_call = :do_not_call,
                notes = :notes, updated_at = now()
            WHERE company_id = CAST(:company_id AS uuid)
              AND (twilio_call_sid = :call_sid OR twilio_parent_call_sid = :call_sid)
            RETURNING *
        """), {**params, "call_sid": call_sid})
        row = existing.first()
    if row is None:
        result = await db.execute(text("""
            INSERT INTO transport_call_logs (
                company_id, advisor_name, advisor_status, customer_name, customer_type, phone,
                origin, destination, trip_type, call_direction, call_status, result, duration_seconds,
                quote_requested, ticket_requested, contract_code, source, twilio_call_sid,
                twilio_parent_call_sid, batch_row_id, campaign_code, caller_number, phone_type,
                price_amount, price_currency, consent_status, do_not_call, notes, created_at, updated_at
            ) VALUES (
                CAST(:company_id AS uuid), :advisor_name, :advisor_status, :customer_name, :customer_type, :phone,
                :origin, :destination, :trip_type, :call_direction, :call_status, :result, :duration_seconds,
                :quote_requested, :ticket_requested, :contract_code, :source, :twilio_call_sid,
                :twilio_parent_call_sid, CAST(NULLIF(:batch_row_id, '') AS uuid), :campaign_code, :caller_number,
                :phone_type, :price_amount, :price_currency, :consent_status, :do_not_call, :notes, now(), now()
            ) RETURNING *
        """), params)
        row = result.first()
    if params["batch_row_id"]:
        await db.execute(
            text(
                """
                UPDATE transport_call_batch_rows
                SET customer_name = COALESCE(NULLIF(:customer_name, ''), customer_name),
                    customer_type = COALESCE(NULLIF(:customer_type, ''), customer_type),
                    phone = COALESCE(NULLIF(:phone, ''), phone),
                    contract_code = COALESCE(NULLIF(:contract_code, ''), contract_code),
                    origin = COALESCE(NULLIF(:origin, ''), origin),
                    destination = COALESCE(NULLIF(:destination, ''), destination),
                    trip_type = COALESCE(NULLIF(:trip_type, ''), trip_type),
                    call_direction = :call_direction,
                    call_status = :call_status,
                    result = :result,
                    duration_seconds = :duration_seconds,
                    quote_requested = :quote_requested,
                    ticket_requested = :ticket_requested,
                    twilio_call_sid = :twilio_call_sid,
                    campaign_code = COALESCE(NULLIF(:campaign_code, ''), campaign_code),
                    phone_type = :phone_type,
                    consent_status = :consent_status,
                    do_not_call = :do_not_call,
                    notes = :notes,
                    managed_by = :advisor_name,
                    managed_at = now(),
                    updated_at = now()
                WHERE id = CAST(:batch_row_id AS uuid)
                  AND company_id = CAST(:company_id AS uuid)
                """
            ),
            {
                "company_id": str(company_id),
                "batch_row_id": params["batch_row_id"],
                "customer_name": _clean(payload.customer_name, 180),
                "customer_type": _clean(payload.customer_type or "person", 40),
                "phone": _clean(payload.phone, 80),
                "contract_code": _clean(payload.contract_code, 120),
                "origin": _clean(payload.origin, 160),
                "destination": _clean(payload.destination, 160),
                "trip_type": _clean(payload.trip_type, 80),
                "call_direction": _clean(payload.call_direction or "inbound", 20),
                "call_status": _clean(payload.call_status or "completed", 40),
                "result": _clean(payload.result or "follow_up", 80),
                "duration_seconds": _duration_seconds(payload),
                "quote_requested": bool(payload.quote_requested),
                "ticket_requested": bool(payload.ticket_requested),
                "twilio_call_sid": _clean(payload.twilio_call_sid or payload.twilio_parent_call_sid, 120),
                "campaign_code": _clean(payload.campaign_code, 120),
                "phone_type": _clean(payload.phone_type or "unknown", 30),
                "consent_status": _clean(payload.consent_status or "unknown", 30),
                "do_not_call": bool(payload.do_not_call),
                "notes": _clean(payload.notes, 1200),
                "advisor_name": _clean(payload.advisor_name, 180),
            },
        )
    await db.commit()
    return {"ok": True, "company_id": str(company_id), "call": _row(row)}


@router.get("/companies/{company_id}/customers", dependencies=[Depends(require_transport_calls_read)])
async def transport_customer_suggestions(
    company_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    search: str = Query(default="", max_length=120),
    employee_id: str = Query(default="", max_length=80),
    limit: int = Query(default=30, ge=1, le=100),
) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    query = f"%{_clean(search, 120).lower()}%"
    if _clean(employee_id, 80):
        assigned = await db.execute(
            text(
                """
                SELECT r.*, b.file_name AS batch_file_name, b.assigned_agent_name
                FROM transport_call_batch_rows r
                JOIN transport_call_batches b ON b.id = r.batch_id
                WHERE r.company_id = CAST(:company_id AS uuid)
                  AND b.status = 'active'
                  AND b.assigned_employee_id = CAST(:employee_id AS uuid)
                  AND r.managed_at IS NULL
                  AND (
                    :query = '%%'
                    OR LOWER(r.customer_name) LIKE :query
                    OR LOWER(r.phone) LIKE :query
                    OR LOWER(r.contract_code) LIKE :query
                    OR LOWER(r.origin) LIKE :query
                    OR LOWER(r.destination) LIKE :query
                  )
                ORDER BY b.created_at DESC, r.row_number ASC
                LIMIT :limit
                """
            ),
            {"company_id": str(company_id), "employee_id": _clean(employee_id, 80), "query": query, "limit": int(limit)},
        )
        rows = []
        for row in assigned.fetchall():
            data = _row(row)
            rows.append({
                "batch_row_id": data.get("id", ""),
                "batch_id": data.get("batch_id", ""),
                "batch_file_name": data.get("batch_file_name", ""),
                "row_number": data.get("row_number", 0),
                "source": "assigned_batch",
                "customer_name": data.get("customer_name", ""),
                "customer_type": data.get("customer_type", "person"),
                "phone": data.get("phone", ""),
                "email": data.get("email", ""),
                "document_id": data.get("document_id", ""),
                "contract_code": data.get("contract_code", ""),
                "account_code": data.get("account_code", ""),
                "origin": data.get("origin", ""),
                "destination": data.get("destination", ""),
                "trip_type": data.get("trip_type", ""),
                "transporter": data.get("transporter", ""),
                "ticket_value": float(data.get("ticket_value") or 0),
                "campaign_code": data.get("campaign_code", ""),
                "phone_type": data.get("phone_type", "unknown"),
                "consent_status": data.get("consent_status", "unknown"),
                "do_not_call": bool(data.get("do_not_call")),
                "notes": data.get("notes", ""),
            })
        return {"ok": True, "company_id": str(company_id), "customers": rows, "count": len(rows)}
    result = await db.execute(
        text(
            """
            SELECT DISTINCT ON (
                NULLIF(LOWER(TRIM(phone)), ''),
                NULLIF(LOWER(TRIM(customer_name)), ''),
                NULLIF(LOWER(TRIM(contract_code)), '')
            )
                customer_name,
                customer_type,
                phone,
                contract_code,
                '' AS account_code,
                origin,
                destination,
                '' AS transporter,
                0 AS ticket_value,
                campaign_code,
                phone_type,
                consent_status,
                do_not_call,
                created_at
            FROM transport_call_logs
            WHERE company_id = :company_id
              AND (
                :query = '%%'
                OR LOWER(customer_name) LIKE :query
                OR LOWER(phone) LIKE :query
                OR LOWER(contract_code) LIKE :query
              )
              AND (
                TRIM(customer_name) <> ''
                OR TRIM(phone) <> ''
                OR TRIM(contract_code) <> ''
              )
            ORDER BY
                NULLIF(LOWER(TRIM(phone)), ''),
                NULLIF(LOWER(TRIM(customer_name)), ''),
                NULLIF(LOWER(TRIM(contract_code)), ''),
                created_at DESC
            LIMIT :limit
            """
        ),
        {"company_id": str(company_id), "query": query, "limit": int(limit)},
    )
    rows = [_row(row) for row in result.fetchall()]
    return {"ok": True, "company_id": str(company_id), "customers": rows, "count": len(rows)}


@router.get("/companies/{company_id}/agents", dependencies=[Depends(require_transport_calls_manage)])
async def list_transport_call_agents(company_id: uuid.UUID, db: AsyncSession = Depends(get_db)) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    work_sessions_available = bool(
        (
            await db.execute(
                text("SELECT to_regclass('mini_panel_work_sessions') IS NOT NULL")
            )
        ).scalar()
    )
    session_fields = """
                   , '' AS work_session_status
                   , 0::bigint AS active_seconds
                   , 0::bigint AS break_seconds
                   , NULL::timestamptz AS work_session_updated_at
    """
    session_join = ""
    if work_sessions_available:
        session_fields = """
                   , COALESCE(ws.status, '') AS work_session_status
                   , (
                       COALESCE(ws.active_seconds, 0)
                       + CASE
                           WHEN ws.status = 'active' AND ws.active_started_at IS NOT NULL
                           THEN GREATEST(0, EXTRACT(EPOCH FROM (now() - ws.active_started_at)))::bigint
                           ELSE 0
                         END
                     )::bigint AS active_seconds
                   , (
                       COALESCE(ws.break_seconds, 0)
                       + CASE
                           WHEN ws.status = 'break' AND ws.current_break_started_at IS NOT NULL
                           THEN GREATEST(0, EXTRACT(EPOCH FROM (now() - ws.current_break_started_at)))::bigint
                           ELSE 0
                         END
                     )::bigint AS break_seconds
                   , ws.updated_at AS work_session_updated_at
        """
        session_join = """
            LEFT JOIN LATERAL (
                SELECT s.status, s.active_seconds, s.break_seconds,
                       s.active_started_at, s.current_break_started_at, s.updated_at
                FROM mini_panel_work_sessions s
                WHERE s.company_id = e.company_id
                  AND s.user_id = u.id
                  AND s.panel_type IN ('call_center', 'external')
                  AND s.status IN ('active', 'break')
                ORDER BY s.started_at DESC
                LIMIT 1
            ) ws ON TRUE
        """
    result = await db.execute(
        text(
            f"""
            SELECT e.id::text AS id, e.full_name, e.phone, e.email, e.role, e.employee_type, e.status,
                   u.id::text AS user_id,
                   COALESCE(u.email, '') AS username,
                   COALESCE(u.role, '') AS user_role,
                   COALESCE(u.settings_json->'mini_panel'->>'type', '') AS mini_panel_type,
                   COALESCE(last_call.advisor_status, '') AS latest_advisor_status,
                   COALESCE(last_call.call_status, '') AS latest_call_status
                   {session_fields}
            FROM employees e
            LEFT JOIN company_users u ON u.company_id = e.company_id
              AND u.settings_json->'mini_panel'->>'employee_id' = e.id::text
              AND u.status = 'active'
            {session_join}
            LEFT JOIN LATERAL (
                SELECT c.advisor_status, c.call_status
                FROM transport_call_logs c
                WHERE c.company_id = e.company_id
                  AND LOWER(TRIM(c.advisor_name)) = LOWER(TRIM(e.full_name))
                ORDER BY c.created_at DESC
                LIMIT 1
            ) last_call ON TRUE
            WHERE e.company_id = CAST(:company_id AS uuid)
              AND LOWER(COALESCE(e.status, 'active')) <> 'archived'
            ORDER BY e.full_name ASC
            """
        ),
        {"company_id": str(company_id)},
    )
    agents = [_row(row) for row in result.fetchall()]
    for agent in agents:
        if not agent.get("role") and agent.get("user_role"):
            agent["role"] = agent.get("user_role")
        session_status = _clean(agent.get("work_session_status"), 40).lower()
        latest_call_status = _clean(agent.get("latest_call_status"), 40).lower()
        latest_advisor_status = _clean(agent.get("latest_advisor_status"), 40).lower()
        if session_status == "break":
            agent["live_status"] = "break"
        elif session_status == "active" and (
            latest_call_status == "pending" or latest_advisor_status == "in_call"
        ):
            agent["live_status"] = "in_call"
        elif session_status == "active":
            agent["live_status"] = "available"
        else:
            agent["live_status"] = "offline"
    agents = [
        agent for agent in agents
        if _is_call_agent_role(agent.get("role") or agent.get("user_role"), agent.get("employee_type"), agent.get("mini_panel_type"))
    ]
    return {"ok": True, "company_id": str(company_id), "agents": agents, "count": len(agents)}


@router.get("/companies/{company_id}/batches", dependencies=[Depends(require_transport_calls_manage)])
async def list_transport_call_batches(
    company_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    status_filter: str = Query(default="active", alias="status", max_length=40),
    limit: int = Query(default=80, ge=1, le=200),
) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    result = await db.execute(
        text(
            """
            SELECT b.*,
                   COUNT(r.id)::int AS rows_total,
                   COUNT(r.id) FILTER (WHERE r.managed_at IS NOT NULL)::int AS rows_managed
            FROM transport_call_batches b
            LEFT JOIN transport_call_batch_rows r ON r.batch_id = b.id
            WHERE b.company_id = CAST(:company_id AS uuid)
              AND (:status = 'all' OR b.status = :status)
            GROUP BY b.id
            ORDER BY b.created_at DESC
            LIMIT :limit
            """
        ),
        {"company_id": str(company_id), "status": _clean(status_filter or "active", 40), "limit": int(limit)},
    )
    rows = [_row(row) for row in result.fetchall()]
    return {"ok": True, "company_id": str(company_id), "batches": rows, "count": len(rows)}


@router.post("/companies/{company_id}/batches/import-csv", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_transport_calls_manage)])
async def import_transport_call_batch_csv(
    company_id: uuid.UUID,
    assigned_employee_id: str = Form(...),
    campaign_code: str = Form(default=""),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    employee = await db.execute(
        text(
            """
            SELECT e.id::text, e.full_name, e.role, e.employee_type,
                   COALESCE(u.role, '') AS user_role,
                   COALESCE(u.settings_json->'mini_panel'->>'type', '') AS mini_panel_type
            FROM employees e
            LEFT JOIN company_users u ON u.company_id = e.company_id
              AND u.settings_json->'mini_panel'->>'employee_id' = e.id::text
              AND u.status = 'active'
            WHERE e.company_id = CAST(:company_id AS uuid)
              AND e.id = CAST(:employee_id AS uuid)
            LIMIT 1
            """
        ),
        {"company_id": str(company_id), "employee_id": _clean(assigned_employee_id, 80)},
    )
    employee_row = _row(employee.first() or {})
    if not employee_row:
        raise HTTPException(status_code=404, detail="agent_not_found")
    if not _is_call_agent_role(
        employee_row.get("role") or employee_row.get("user_role"),
        employee_row.get("employee_type"),
        employee_row.get("mini_panel_type"),
    ):
        raise HTTPException(status_code=400, detail="employee_is_not_call_agent")
    active_batch = await db.execute(
        text(
            """
            SELECT id::text
            FROM transport_call_batches
            WHERE company_id = CAST(:company_id AS uuid)
              AND assigned_employee_id = CAST(:employee_id AS uuid)
              AND status = 'active'
            LIMIT 1
            """
        ),
        {"company_id": str(company_id), "employee_id": _clean(assigned_employee_id, 80)},
    )
    if active_batch.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="agent_has_active_batch")
    user = await db.execute(
        text(
            """
            SELECT id::text
            FROM company_users
            WHERE company_id = CAST(:company_id AS uuid)
              AND settings_json->'mini_panel'->>'employee_id' = :employee_id
              AND status = 'active'
            LIMIT 1
            """
        ),
        {"company_id": str(company_id), "employee_id": _clean(assigned_employee_id, 80)},
    )
    user_id = user.scalar_one_or_none() or ""
    filename = _clean(file.filename or "base_llamadas.csv", 240)
    raw = await file.read()
    try:
        rows = _uploaded_call_rows(raw, filename)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    campaign = _clean(campaign_code, 120) or _clean(filename.rsplit(".", 1)[0].replace("_", " "), 120) or "General"
    batch = await db.execute(
        text(
            """
            INSERT INTO transport_call_batches (
                company_id, file_name, assigned_employee_id, assigned_agent_name,
                assigned_agent_role, assigned_user_id, record_count, campaign_code, created_by_label
            ) VALUES (
                CAST(:company_id AS uuid), :file_name, CAST(:employee_id AS uuid), :agent_name,
                :agent_role, CAST(NULLIF(:user_id, '') AS uuid), :record_count, :campaign_code, 'Admin V2'
            )
            RETURNING id
            """
        ),
        {
            "company_id": str(company_id),
            "file_name": filename,
            "employee_id": _clean(assigned_employee_id, 80),
            "agent_name": _clean(employee_row.get("full_name"), 180),
            "agent_role": _clean(employee_row.get("role") or employee_row.get("employee_type"), 80),
            "user_id": _clean(user_id, 80),
            "record_count": len(rows),
            "campaign_code": campaign,
        },
    )
    batch_id = str(batch.scalar_one())
    for index, row in enumerate(rows, start=1):
        await db.execute(
            text(
                """
                INSERT INTO transport_call_batch_rows (
                    batch_id, company_id, row_number, customer_name, customer_type, phone, email,
                    document_id, contract_code, account_code, origin, destination, trip_type,
                    transporter, ticket_value, campaign_code, phone_type, consent_status, do_not_call, notes
                ) VALUES (
                    CAST(:batch_id AS uuid), CAST(:company_id AS uuid), :row_number, :customer_name, :customer_type, :phone, :email,
                    :document_id, :contract_code, :account_code, :origin, :destination, :trip_type,
                    :transporter, :ticket_value, :campaign_code, :phone_type, :consent_status, :do_not_call, :notes
                )
                """
            ),
            {
                "batch_id": batch_id,
                "company_id": str(company_id),
                "row_number": index,
                "customer_name": _csv_value(row, "cliente", "customer_name", "nombre", "empresa", "razon_social", limit=180),
                "customer_type": _csv_value(row, "tipo_cliente", "customer_type", limit=40) or "person",
                "phone": _csv_value(row, "telefono", "phone", "celular", "whatsapp", limit=80),
                "email": _csv_value(row, "correo", "email", "mail", limit=160),
                "document_id": _csv_value(row, "documento", "nit", "cc", "id", limit=80),
                "contract_code": _csv_value(row, "contrato", "aval", "contrato_aval", "numero_contrato", limit=120),
                "account_code": _csv_value(row, "cuenta", "account_code", "cuenta_k", "codigo_cuenta", limit=120),
                "origin": _csv_value(row, "origen", "ciudad_origen", limit=160),
                "destination": _csv_value(row, "destino", "ciudad_destino", limit=160),
                "trip_type": _csv_value(row, "tipo_viaje", "servicio", "ruta", limit=80),
                "transporter": _csv_value(row, "transportadora", "empresa_transportadora", "operador", "carrier", limit=180),
                "ticket_value": _csv_money(row, "valor_ticket", "valor", "tarifa", "precio", "valor_base", "unit_value", "valor_unitario"),
                "campaign_code": _csv_value(row, "campana", "campaign", "campaign_code", limit=120) or campaign,
                "phone_type": _csv_value(row, "tipo_linea", "phone_type", limit=30) or "unknown",
                "consent_status": _csv_value(row, "consentimiento", "consent_status", limit=30) or "unknown",
                "do_not_call": _csv_bool(row, "no_llamar", "do_not_call", "lista_robinson"),
                "notes": _csv_value(row, "notas", "observaciones", "comentarios", limit=1200),
            },
        )
    await db.commit()
    return {"ok": True, "company_id": str(company_id), "batch_id": batch_id, "record_count": len(rows)}


@router.post("/companies/{company_id}/batches/{batch_id}/archive", dependencies=[Depends(require_transport_calls_manage)])
async def archive_transport_call_batch(company_id: uuid.UUID, batch_id: uuid.UUID, db: AsyncSession = Depends(get_db)) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    await db.execute(
        text(
            """
            UPDATE transport_call_batches
            SET status = 'archived', archived_at = now(), updated_at = now()
            WHERE company_id = CAST(:company_id AS uuid)
              AND id = CAST(:batch_id AS uuid)
            """
        ),
        {"company_id": str(company_id), "batch_id": str(batch_id)},
    )
    await db.commit()
    return {"ok": True, "company_id": str(company_id), "batch_id": str(batch_id)}


@router.delete("/companies/{company_id}/batches/{batch_id}", dependencies=[Depends(require_transport_calls_manage)])
async def delete_transport_call_batch(company_id: uuid.UUID, batch_id: uuid.UUID, db: AsyncSession = Depends(get_db)) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    await db.execute(
        text(
            """
            DELETE FROM transport_call_batches
            WHERE company_id = CAST(:company_id AS uuid)
              AND id = CAST(:batch_id AS uuid)
            """
        ),
        {"company_id": str(company_id), "batch_id": str(batch_id)},
    )
    await db.commit()
    return {"ok": True, "company_id": str(company_id), "batch_id": str(batch_id)}


@router.get("/companies/{company_id}/batches/{batch_id}/export.csv", dependencies=[Depends(require_transport_calls_manage)])
async def export_transport_call_batch_csv(company_id: uuid.UUID, batch_id: uuid.UUID, db: AsyncSession = Depends(get_db)) -> Response:
    await ensure_transport_calls_storage(db)
    result = await db.execute(
        text(
            """
            SELECT b.file_name, b.assigned_agent_name, r.*,
                   COALESCE(call_cost.price_amount, r.price_amount, 0) AS exported_price_amount,
                   COALESCE(call_cost.price_currency, r.price_currency, 'USD') AS exported_price_currency
            FROM transport_call_batches b
            JOIN transport_call_batch_rows r ON r.batch_id = b.id
            LEFT JOIN LATERAL (
                SELECT c.price_amount, c.price_currency
                FROM transport_call_logs c
                WHERE c.company_id = r.company_id AND c.batch_row_id = r.id
                ORDER BY c.updated_at DESC
                LIMIT 1
            ) call_cost ON TRUE
            WHERE b.company_id = CAST(:company_id AS uuid)
              AND b.id = CAST(:batch_id AS uuid)
            ORDER BY r.row_number ASC
            """
        ),
        {"company_id": str(company_id), "batch_id": str(batch_id)},
    )
    rows = [_row(row) for row in result.fetchall()]
    output = io.StringIO()
    fieldnames = [
        "archivo", "agente", "fila", "cliente", "tipo_cliente", "telefono", "correo", "documento", "contrato",
        "cuenta", "campana", "tipo_linea", "consentimiento", "no_llamar", "origen", "destino", "tipo_viaje", "transportadora", "valor_ticket",
        "direccion_llamada", "estado_llamada", "resultado", "duracion_segundos",
        "genero_cotizacion", "genero_ticket", "id_twilio", "costo_llamada", "moneda", "notas", "gestionado_por", "gestionado_en",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow({
            "archivo": row.get("file_name") or "",
            "agente": row.get("assigned_agent_name") or "",
            "fila": row.get("row_number") or "",
            "cliente": row.get("customer_name") or "",
            "tipo_cliente": row.get("customer_type") or "",
            "telefono": row.get("phone") or "",
            "correo": row.get("email") or "",
            "documento": row.get("document_id") or "",
            "contrato": row.get("contract_code") or "",
            "cuenta": row.get("account_code") or "",
            "campana": row.get("campaign_code") or "",
            "tipo_linea": row.get("phone_type") or "unknown",
            "consentimiento": row.get("consent_status") or "unknown",
            "no_llamar": "si" if row.get("do_not_call") else "no",
            "origen": row.get("origin") or "",
            "destino": row.get("destination") or "",
            "tipo_viaje": row.get("trip_type") or "",
            "transportadora": row.get("transporter") or "",
            "valor_ticket": row.get("ticket_value") or 0,
            "direccion_llamada": row.get("call_direction") or "",
            "estado_llamada": row.get("call_status") or "",
            "resultado": row.get("result") or "",
            "duracion_segundos": row.get("duration_seconds") or 0,
            "genero_cotizacion": "si" if row.get("quote_requested") else "no",
            "genero_ticket": "si" if row.get("ticket_requested") else "no",
            "id_twilio": row.get("twilio_call_sid") or "",
            "costo_llamada": row.get("exported_price_amount") or 0,
            "moneda": row.get("exported_price_currency") or "USD",
            "notas": row.get("notes") or "",
            "gestionado_por": row.get("managed_by") or "",
            "gestionado_en": row.get("managed_at") or "",
        })
    filename = "gestion_llamadas.csv"
    if rows and rows[0].get("file_name"):
        filename = f"gestion_{str(rows[0]['file_name']).replace(' ', '_')}"
        if not filename.lower().endswith(".csv"):
            filename += ".csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _twilio_direction(value: Any) -> str:
    raw = _clean(value, 40).lower()
    if "out" in raw:
        return "outbound"
    return "inbound"


def _twilio_status(value: Any) -> str:
    raw = _clean(value, 40).lower()
    if raw in {"no-answer", "busy", "failed", "canceled", "cancelled"}:
        return "missed"
    if raw in {"ringing", "queued", "initiated", "in-progress"}:
        return "pending"
    return "completed"


@router.post("/companies/{company_id}/twilio/status")
async def transport_twilio_status_webhook(
    company_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    await ensure_transport_calls_storage(db)
    if "application/json" in _clean(request.headers.get("content-type"), 120).lower():
        form = await request.json()
    else:
        raw_body = (await request.body()).decode("utf-8", errors="ignore")
        parsed = parse_qs(raw_body)
        form = {key: values[-1] if values else "" for key, values in parsed.items()}

    call_sid = _clean(form.get("CallSid") or form.get("call_sid"), 120)
    if not call_sid:
        return {"ok": False, "detail": "missing_call_sid"}

    direction = _twilio_direction(form.get("Direction"))
    status_value = _twilio_status(form.get("CallStatus"))
    duration = max(0, int(float(form.get("CallDuration") or form.get("Duration") or 0)))
    phone = _clean(form.get("From") or form.get("Caller") or "", 80)
    notes = _clean(form.get("To") or "", 1200)

    existing = await db.execute(
        text(
            """
            SELECT id
            FROM transport_call_logs
            WHERE company_id = :company_id
              AND twilio_call_sid = :call_sid
            LIMIT 1
            """
        ),
        {"company_id": str(company_id), "call_sid": call_sid},
    )
    row = existing.first()

    if row:
        result = await db.execute(
            text(
                """
                UPDATE transport_call_logs
                SET call_direction = :direction,
                    call_status = :status,
                    duration_seconds = CASE WHEN :duration > 0 THEN :duration ELSE duration_seconds END,
                    phone = CASE WHEN TRIM(phone) = '' THEN :phone ELSE phone END,
                    source = 'twilio',
                    updated_at = now()
                WHERE id = :id
                RETURNING *
                """
            ),
            {
                "id": row._mapping["id"],
                "direction": direction,
                "status": status_value,
                "duration": duration,
                "phone": phone,
            },
        )
    else:
        result = await db.execute(
            text(
                """
                INSERT INTO transport_call_logs (
                    company_id,
                    advisor_status,
                    customer_type,
                    phone,
                    call_direction,
                    call_status,
                    duration_seconds,
                    source,
                    twilio_call_sid,
                    notes,
                    created_at,
                    updated_at
                )
                VALUES (
                    :company_id,
                    :advisor_status,
                    'person',
                    :phone,
                    :direction,
                    :status,
                    :duration,
                    'twilio',
                    :call_sid,
                    :notes,
                    now(),
                    now()
                )
                RETURNING *
                """
            ),
            {
                "company_id": str(company_id),
                "advisor_status": "in_call" if status_value == "pending" else "available",
                "phone": phone,
                "direction": direction,
                "status": status_value,
                "duration": duration,
                "call_sid": call_sid,
                "notes": notes,
            },
        )

    await db.commit()
    return {"ok": True, "company_id": str(company_id), "call": _row(result.first())}
